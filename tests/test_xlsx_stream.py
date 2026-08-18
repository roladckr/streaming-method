import gc
import tempfile
import time
import unittest
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook

from zipfile import ZipFile

from src.streaming.xlsx_stream import (
    MAIN_NS,
    _iter_shared_strings,
    load_shared_strings,
    read_header_row,
    stream_xlsx_rows,
)

FIXTURE = Path(__file__).parent / "fixtures" / "sample_c2.xlsx"
VCC_FIXTURE = (
    Path(__file__).parent / "fixtures" / "sample_vcc_step_08.xlsx"
)


class GenericStreamingTests(unittest.TestCase):
    """
    Confirms the streaming engine reads arbitrary columns generically -
    including column B, which no processor (C2 or otherwise) reads.
    """

    def test_reads_requested_columns_for_every_row(self):
        rows = list(
            stream_xlsx_rows(
                FIXTURE,
                requested_columns=["A", "B"],
            )
        )

        self.assertEqual(
            [(row.row_number, row.values) for row in rows],
            [
                (1, {"A": "Item", "B": "Extra"}),
                (2, {"A": "SYN-1001", "B": "alpha"}),
                (3, {"A": "SYN-1002", "B": "bravo"}),
                (4, {"A": "SYN-1001", "B": "charlie"}),
                (5, {"A": "SYN-1003", "B": "delta"}),
            ],
        )

    def test_columns_not_requested_are_not_returned(self):
        rows = list(
            stream_xlsx_rows(
                FIXTURE,
                requested_columns=["A"],
            )
        )

        for row in rows:
            self.assertEqual(set(row.values.keys()), {"A"})


class ReadHeaderRowTests(unittest.TestCase):
    """
    Generic capability: resolving column letters by header name,
    for processors (like VCC_STEP_08) that don't have a verified
    fixed column layout the way CAR_RENTAL_C2 does.
    """

    def test_maps_header_text_to_column_letters(self):
        headers = read_header_row(FIXTURE)

        self.assertEqual(
            headers,
            {
                "Item": "A",
                "Extra": "B",
                "VAR": "D",
                "Order Status": "G",
                "Est Total Price": "AR",
            },
        )

    def test_works_with_headers_in_different_column_positions(self):
        # sample_vcc_step_08.xlsx deliberately uses a different
        # column layout than sample_c2.xlsx, to prove this isn't
        # assuming any fixed positions.
        headers = read_header_row(VCC_FIXTURE)

        self.assertEqual(headers["Order ID"], "B")
        self.assertEqual(headers["Line Num"], "C")
        self.assertEqual(headers["Retail Price"], "E")
        self.assertEqual(
            headers["bridge2Unit Base Price"], "F"
        )
        self.assertEqual(headers["Gateway ID"], "H")

    def test_stops_after_header_row_on_a_large_file(self):
        # Resolving headers on a large file (e.g. a 400K-row DAC
        # export) must not require reading the whole file. Build a
        # sheet where only row 1 has proper headers and prove
        # read_header_row still returns correctly and quickly.
        with tempfile.TemporaryDirectory() as tmpdir:
            large_fixture = Path(tmpdir) / "large.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.cell(row=1, column=1, value="Order ID")
            for row_index in range(2, 5001):
                worksheet.cell(
                    row=row_index,
                    column=1,
                    value=f"SYN-{row_index}",
                )
            workbook.save(large_fixture)

            start = time.monotonic()
            headers = read_header_row(large_fixture)
            elapsed = time.monotonic() - start

        self.assertEqual(headers, {"Order ID": "A"})
        self.assertLess(elapsed, 1.0)


class SheetDataMemoryRegressionTests(unittest.TestCase):
    """
    Regression test for the sheetData retention fix.

    On the pre-fix implementation, every processed <row> element stayed
    attached to sheetData for the life of the parse, so the number of
    live row elements reachable mid-stream grows in lockstep with rows
    already consumed. After the fix, sheetData only ever holds a small,
    bounded number of rows (parser look-ahead), independent of how far
    into the file the generator has read.

    Uses a larger on-the-fly fixture (not committed - see
    tests/fixtures/ which is reserved for small, committed fixtures)
    so the "grows with rows consumed" signature is unambiguous.
    """

    ROW_COUNT = 300
    LEAK_BOUND = 20  # comfortably above normal parser look-ahead noise

    @classmethod
    def setUpClass(cls):
        cls._tmpdir = tempfile.TemporaryDirectory()
        cls.large_fixture = (
            Path(cls._tmpdir.name) / "large_synthetic.xlsx"
        )

        workbook = Workbook()
        worksheet = workbook.active
        for row_index in range(1, cls.ROW_COUNT + 1):
            worksheet.cell(
                row=row_index,
                column=1,
                value=f"SYN-{row_index}",
            )
        workbook.save(cls.large_fixture)

    @classmethod
    def tearDownClass(cls):
        cls._tmpdir.cleanup()

    def test_row_retention_does_not_scale_with_rows_consumed(self):
        row_tag = f"{{{MAIN_NS}}}row"

        generator = stream_xlsx_rows(
            self.large_fixture,
            requested_columns=["A"],
        )

        consume_count = self.ROW_COUNT - 10
        for _ in range(consume_count):
            next(generator)

        gc.collect()

        leaked_rows = [
            obj
            for obj in gc.get_objects()
            if isinstance(obj, ET.Element) and obj.tag == row_tag
        ]

        self.assertLess(
            len(leaked_rows),
            self.LEAK_BOUND,
            f"{len(leaked_rows)} row elements still reachable after "
            f"consuming {consume_count}/{self.ROW_COUNT} rows - "
            "sheetData appears to be retaining processed rows",
        )

        # Drain the remainder so the underlying file handle closes
        # cleanly instead of relying on generator GC.
        list(generator)


class SharedStringsMemoryRegressionTests(unittest.TestCase):
    """
    Regression test for the sharedStrings.xml retention fix.

    On the pre-fix implementation, every decoded <si> element stayed
    attached to the <sst> root for the life of the parse (cleared, but
    never detached), so the number of live si elements reachable
    mid-parse grows in lockstep with entries already decoded. After
    the fix, the root only ever holds a small, bounded number of si
    elements, independent of how many shared strings have already
    been consumed.

    Uses a larger on-the-fly fixture with many unique strings (shared
    strings only exist per-unique-value, so this requires distinct
    text per row) so the "grows with entries consumed" signature is
    unambiguous.
    """

    STRING_COUNT = 300
    LEAK_BOUND = 20  # comfortably above normal parser look-ahead noise

    @classmethod
    def setUpClass(cls):
        # openpyxl always writes text cells as inline strings
        # (t="inlineStr") on save, never via xl/sharedStrings.xml -
        # confirmed against both openpyxl's own output and the
        # existing committed fixtures, neither of which contain a
        # sharedStrings part. Real Excel/Drive-exported files
        # routinely do use shared strings, so this fixture is built
        # by hand (zip + the minimal sharedStrings.xml part) to
        # actually exercise that code path.
        cls._tmpdir = tempfile.TemporaryDirectory()
        cls.large_fixture = (
            Path(cls._tmpdir.name) / "many_shared_strings.xlsx"
        )
        cls.expected_strings = [
            f"UNIQUE-STRING-{index}"
            for index in range(cls.STRING_COUNT)
        ]

        shared_strings_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<sst xmlns="{MAIN_NS}" '
            f'count="{cls.STRING_COUNT}" '
            f'uniqueCount="{cls.STRING_COUNT}">'
            + "".join(
                f"<si><t>{value}</t></si>"
                for value in cls.expected_strings
            )
            + "</sst>"
        )

        with ZipFile(cls.large_fixture, "w") as archive:
            archive.writestr(
                "xl/sharedStrings.xml", shared_strings_xml
            )

    @classmethod
    def tearDownClass(cls):
        cls._tmpdir.cleanup()

    def test_decoded_values_are_exact(self):
        # Exercises the actual public entry point used by
        # stream_xlsx_rows/read_header_row, not just the internal
        # generator.
        with ZipFile(self.large_fixture) as archive:
            strings = load_shared_strings(archive)

        self.assertEqual(strings, self.expected_strings)

    def test_si_retention_does_not_scale_with_entries_consumed(self):
        si_tag = f"{{{MAIN_NS}}}si"

        with ZipFile(self.large_fixture) as archive:
            with archive.open(
                "xl/sharedStrings.xml"
            ) as source:
                generator = _iter_shared_strings(source)

                consume_count = self.STRING_COUNT - 10
                for _ in range(consume_count):
                    next(generator)

                gc.collect()

                leaked_entries = [
                    obj
                    for obj in gc.get_objects()
                    if isinstance(obj, ET.Element)
                    and obj.tag == si_tag
                ]

                self.assertLess(
                    len(leaked_entries),
                    self.LEAK_BOUND,
                    f"{len(leaked_entries)} si elements still "
                    f"reachable after consuming "
                    f"{consume_count}/{self.STRING_COUNT} entries - "
                    "sharedStrings root appears to be retaining "
                    "processed entries",
                )

                # Drain the remainder so the underlying file handle
                # closes cleanly instead of relying on generator GC.
                list(generator)


if __name__ == "__main__":
    unittest.main()

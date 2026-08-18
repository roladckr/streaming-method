import gc
import tempfile
import time
import unittest
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook

from src.processors.vcc_step_08 import (
    process_vcc_step_08,
    resolve_dac_columns,
)
from src.streaming.xlsx_stream import MAIN_NS, stream_xlsx_rows

FIXTURE = (
    Path(__file__).parent / "fixtures" / "sample_vcc_step_08.xlsx"
)


class VccStep08LookupTests(unittest.TestCase):
    def test_one_order_id_to_one_row(self):
        result = process_vcc_step_08(
            file_paths=[FIXTURE],
            lookup_keys=["DAC-1003"],
        )

        self.assertEqual(result["matched_order_count"], 1)
        self.assertEqual(result["matched_row_count"], 1)

        match = result["matches"][0]
        self.assertEqual(match["Order ID"], "DAC-1003")
        self.assertEqual(match["Line Num"], 1)
        self.assertEqual(match["Retail Price"], 75)
        self.assertEqual(
            match["bridge2Unit Base Price"], 60
        )
        self.assertEqual(match["Gateway ID"], "GW-CCC")

    def test_one_order_id_to_multiple_line_num_rows(self):
        # DAC-1001 appears twice in the fixture (Line Num 1 and 2).
        # Unlike C2, BOTH rows must be returned - no dedup by
        # Order ID.
        result = process_vcc_step_08(
            file_paths=[FIXTURE],
            lookup_keys=["DAC-1001"],
        )

        self.assertEqual(result["matched_order_count"], 1)
        self.assertEqual(result["matched_row_count"], 2)

        line_nums = sorted(
            match["Line Num"] for match in result["matches"]
        )
        self.assertEqual(line_nums, [1, 2])

        for match in result["matches"]:
            self.assertEqual(match["Order ID"], "DAC-1001")

    def test_multiple_requested_order_ids(self):
        result = process_vcc_step_08(
            file_paths=[FIXTURE],
            lookup_keys=["DAC-1001", "DAC-1002", "DAC-1003"],
        )

        self.assertEqual(result["requested_count"], 3)
        self.assertEqual(result["matched_order_count"], 3)
        # DAC-1001 (x2 lines) + DAC-1002 (x1) + DAC-1003 (x1) = 4
        self.assertEqual(result["matched_row_count"], 4)
        self.assertEqual(result["not_found"], [])

    def test_not_found_order_id_is_reported(self):
        result = process_vcc_step_08(
            file_paths=[FIXTURE],
            lookup_keys=["DAC-1001", "DAC-9999"],
        )

        self.assertEqual(result["requested_count"], 2)
        self.assertEqual(result["matched_order_count"], 1)
        self.assertEqual(result["not_found"], ["DAC-9999"])

    def test_numeric_order_id_normalizes_against_string_lookup(self):
        # Row stores Order ID 555000 as a real Excel number, not
        # text. Both plain-string and ".0"-suffixed JSON lookups
        # must match it.
        for lookup_key in ("555000", "555000.0"):
            with self.subTest(lookup_key=lookup_key):
                result = process_vcc_step_08(
                    file_paths=[FIXTURE],
                    lookup_keys=[lookup_key],
                )

                self.assertEqual(
                    result["matched_row_count"], 1
                )
                self.assertEqual(
                    result["matches"][0]["Order ID"], "555000"
                )
                self.assertEqual(
                    result["matches"][0]["Gateway ID"], "GW-NUM"
                )

    def test_returned_fields_are_complete(self):
        result = process_vcc_step_08(
            file_paths=[FIXTURE],
            lookup_keys=["DAC-1002"],
        )

        match = result["matches"][0]
        self.assertEqual(
            set(match.keys()),
            {
                "Order ID",
                "Line Num",
                "Retail Price",
                "bridge2Unit Base Price",
                "Gateway ID",
                "_source_file",
                "_source_row",
            },
        )


class VccStep08ColumnResolutionTests(unittest.TestCase):
    def test_resolves_columns_by_header_not_position(self):
        # sample_vcc_step_08.xlsx uses different column letters
        # than C2's fixture, proving this isn't a hardcoded
        # positional assumption.
        columns = resolve_dac_columns(FIXTURE)

        self.assertEqual(
            columns,
            {
                "Order ID": "B",
                "Line Num": "C",
                "Retail Price": "E",
                "bridge2Unit Base Price": "F",
                "Gateway ID": "H",
            },
        )

    def test_missing_header_raises_clear_error(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            bad_fixture = Path(tmpdir) / "missing_headers.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.cell(row=1, column=1, value="Order ID")
            # Missing Line Num, Retail Price, bridge2Unit Base
            # Price, Gateway ID.
            workbook.save(bad_fixture)

            with self.assertRaises(ValueError):
                resolve_dac_columns(bad_fixture)


class VccStep08LargeFileTests(unittest.TestCase):
    """
    Confirms VCC_STEP_08 correctly reuses the shared streaming
    engine's memory-bounded behavior on a larger synthetic sheet,
    rather than accidentally materializing the whole file.
    """

    ROW_COUNT = 2000
    LEAK_BOUND = 20
    TARGET_ORDER_ID = "DAC-TARGET"
    MATCH_EVERY = 500

    @classmethod
    def setUpClass(cls):
        cls._tmpdir = tempfile.TemporaryDirectory()
        cls.large_fixture = (
            Path(cls._tmpdir.name) / "large_dac.xlsx"
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=2, value="Order ID")
        worksheet.cell(row=1, column=3, value="Line Num")
        worksheet.cell(row=1, column=5, value="Retail Price")
        worksheet.cell(
            row=1, column=6, value="bridge2Unit Base Price"
        )
        worksheet.cell(row=1, column=8, value="Gateway ID")

        for row_index in range(2, cls.ROW_COUNT + 2):
            order_id = (
                cls.TARGET_ORDER_ID
                if row_index % cls.MATCH_EVERY == 0
                else f"DAC-{row_index}"
            )
            worksheet.cell(row=row_index, column=2, value=order_id)
            worksheet.cell(row=row_index, column=3, value=1)
            worksheet.cell(row=row_index, column=5, value=100.0)
            worksheet.cell(row=row_index, column=6, value=90.0)
            worksheet.cell(row=row_index, column=8, value="GW-X")

        workbook.save(cls.large_fixture)

    @classmethod
    def tearDownClass(cls):
        cls._tmpdir.cleanup()

    def test_large_file_matches_correctly_and_completes_quickly(
        self,
    ):
        start = time.monotonic()
        result = process_vcc_step_08(
            file_paths=[self.large_fixture],
            lookup_keys=[self.TARGET_ORDER_ID],
        )
        elapsed = time.monotonic() - start

        expected_hits = self.ROW_COUNT // self.MATCH_EVERY
        self.assertEqual(
            result["matched_row_count"], expected_hits
        )
        self.assertEqual(result["matched_order_count"], 1)
        self.assertLess(elapsed, 5.0)

    def test_row_retention_stays_bounded_with_dac_columns(self):
        # Same regression technique as
        # tests/test_xlsx_stream.py::SheetDataMemoryRegressionTests,
        # applied specifically to VCC's header-resolved columns, to
        # confirm resolve_dac_columns doesn't change stream_xlsx_rows
        # memory behavior.
        columns = resolve_dac_columns(self.large_fixture)
        row_tag = f"{{{MAIN_NS}}}row"

        generator = stream_xlsx_rows(
            self.large_fixture,
            requested_columns=columns.values(),
        )

        consume_count = self.ROW_COUNT - 10
        for _ in range(consume_count):
            next(generator)

        gc.collect()

        leaked_rows = [
            obj
            for obj in gc.get_objects()
            if isinstance(obj, ET.Element) and obj.tag == row_tag
        ]

        self.assertLess(
            len(leaked_rows),
            self.LEAK_BOUND,
            f"{len(leaked_rows)} row elements still reachable after "
            f"consuming {consume_count}/{self.ROW_COUNT} rows",
        )

        list(generator)


if __name__ == "__main__":
    unittest.main()
